"""
השוואת נסחי טאבו בין שני חודשים — איתור שינויים ויצירת דוח אקסל ידידותי לאישור.

שימוש:
    python tabu-monthly/compare_tabu_months.py "נסחים מרץ 2026" "נסחים אפריל 2026"

פלט:
    דוחות/דוח_שינויים_<חודש_ישן>_ל<חודש_חדש>.xlsx — דוח החלטות (תסמן ידנית)
    assets/tabu_owners.json — מתעדכן לגרסה החדשה (אחרי אישור)
"""
import os
import sys
import re
import json
from glob import glob
from collections import defaultdict

sys.stdout.reconfigure(encoding='utf-8')
BASE_DIR = r'C:\Users\User2\Desktop\סימולטור עסקת קומבינציה וטבלאות ההקצעה'
os.chdir(BASE_DIR)

# שטחי חלקות (להשוואה)
PARCEL_AREAS = {
    3: 2266, 5: 13794, 6: 7764, 7: 5096, 8: 11969, 9: 9126,
    11: 5357, 12: 12232, 13: 302, 14: 14152, 15: 17264, 16: 9022,
    17: 16987, 18: 16830, 19: 26980, 20: 10904, 21: 4573, 22: 155,
    26: 183, 28: 27335, 31: 22270, 32: 22393, 34: 7400, 36: 70,
    39: 15000, 40: 7371, 41: 351, 42: 1807, 44: 13169, 45: 18236,
    46: 5628, 49: 1105, 51: 24918, 53: 12926, 55: 14600, 57: 5327,
    59: 12461, 61: 10000, 66: 7000, 67: 10036, 68: 10032, 69: 10035, 70: 10039,
}


def parse_folder(folder):
    """קורא את כל הנסחים בתיקייה ומפרסר ל-JSON. מחזיר dict."""
    # ייבוא מהמקום הנכון
    sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
    from parse_all_tabu import parse_tabu

    result = {}
    files = sorted(glob(os.path.join(folder, '3852-*.pdf')),
                   key=lambda f: int(re.search(r'3852-(\d+)\.pdf', f).group(1)))

    for f in files:
        m = re.search(r'3852-(\d+)\.pdf', f)
        if not m: continue
        parcel_num = int(m.group(1))
        if parcel_num not in PARCEL_AREAS: continue
        owners = parse_tabu(f, parcel_num)
        result[str(parcel_num)] = {
            'parcel_area': PARCEL_AREAS[parcel_num],
            'owners': owners,
        }
    return result


def compare_owners(old_owners, new_owners):
    """משווה רשימות בעלים בחלקה אחת. מחזיר {added, removed, changed_area}."""
    old_by_id = {o['id']: o for o in old_owners}
    new_by_id = {o['id']: o for o in new_owners}

    old_ids = set(old_by_id.keys())
    new_ids = set(new_by_id.keys())

    added = [new_by_id[i] for i in (new_ids - old_ids)]
    removed = [old_by_id[i] for i in (old_ids - new_ids)]
    changed_area = []
    for i in (old_ids & new_ids):
        old_a = old_by_id[i]['area_sqm']
        new_a = new_by_id[i]['area_sqm']
        if abs(old_a - new_a) > 1:
            changed_area.append({
                'id': i,
                'name': new_by_id[i]['name'],
                'kind': new_by_id[i]['kind'],
                'old_area': old_a,
                'new_area': new_a,
                'diff': new_a - old_a,
            })
    return {'added': added, 'removed': removed, 'changed_area': changed_area}


def build_excel_report(diffs, output_path):
    """בונה אקסל אינטראקטיבי לאישור החלטות."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.worksheet.dimensions import ColumnDimension

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='1A3A5C')
    added_fill = PatternFill('solid', fgColor='D1FAE5')
    removed_fill = PatternFill('solid', fgColor='FEE2E2')
    changed_fill = PatternFill('solid', fgColor='FEF3C7')

    # גיליון 1 — בעלים חדשים
    ws1 = wb.create_sheet('1 - בעלים חדשים')
    ws1.sheet_view.rightToLeft = True
    ws1.append(['חלקה', 'שם', 'תעודה', 'סוג', 'חלק', 'שטח (מ"ר)', 'פעולה', 'החלטה'])
    for cell in ws1[1]:
        cell.font = header_font
        cell.fill = header_fill
    for parcel_str, parcel_diff in sorted(diffs.items(), key=lambda x: int(x[0])):
        for o in parcel_diff['added']:
            row = [int(parcel_str), o['name'], o['id'], o['kind'], o['share_str'],
                   round(o['area_sqm'], 2), o.get('action') or '', '']
            ws1.append(row)
            for cell in ws1[ws1.max_row]:
                cell.fill = added_fill

    # גיליון 2 — בעלים שיצאו
    ws2 = wb.create_sheet('2 - בעלים שיצאו')
    ws2.sheet_view.rightToLeft = True
    ws2.append(['חלקה', 'שם', 'תעודה', 'סוג', 'שטח קודם (מ"ר)', 'החלטה'])
    for cell in ws2[1]:
        cell.font = header_font
        cell.fill = header_fill
    for parcel_str, parcel_diff in sorted(diffs.items(), key=lambda x: int(x[0])):
        for o in parcel_diff['removed']:
            row = [int(parcel_str), o['name'], o['id'], o['kind'],
                   round(o['area_sqm'], 2), '']
            ws2.append(row)
            for cell in ws2[ws2.max_row]:
                cell.fill = removed_fill

    # גיליון 3 — שינויי שטח
    ws3 = wb.create_sheet('3 - שינויי שטח')
    ws3.sheet_view.rightToLeft = True
    ws3.append(['חלקה', 'שם', 'תעודה', 'שטח קודם', 'שטח חדש', 'הפרש (+/-)', 'החלטה'])
    for cell in ws3[1]:
        cell.font = header_font
        cell.fill = header_fill
    for parcel_str, parcel_diff in sorted(diffs.items(), key=lambda x: int(x[0])):
        for c in parcel_diff['changed_area']:
            row = [int(parcel_str), c['name'], c['id'],
                   round(c['old_area'], 2), round(c['new_area'], 2),
                   round(c['diff'], 2), '']
            ws3.append(row)
            for cell in ws3[ws3.max_row]:
                cell.fill = changed_fill

    # גיליון 4 — סיכום
    ws4 = wb.create_sheet('4 - סיכום', 0)
    ws4.sheet_view.rightToLeft = True
    ws4.append(['חלקה', 'בעלים חדשים', 'בעלים שיצאו', 'שינויי שטח', 'סה"כ שינויים'])
    for cell in ws4[1]:
        cell.font = header_font
        cell.fill = header_fill
    for parcel_str, parcel_diff in sorted(diffs.items(), key=lambda x: int(x[0])):
        n_added = len(parcel_diff['added'])
        n_removed = len(parcel_diff['removed'])
        n_changed = len(parcel_diff['changed_area'])
        total = n_added + n_removed + n_changed
        if total == 0: continue
        ws4.append([int(parcel_str), n_added, n_removed, n_changed, total])

    # רוחב עמודות אוטומטי
    for ws in wb.worksheets:
        for col in ws.columns:
            max_len = max((len(str(c.value)) if c.value else 0 for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max(12, max_len + 2), 40)

    wb.save(output_path)
    print(f'✅ נשמר: {output_path}')


def main():
    if len(sys.argv) < 3:
        print('שימוש: python compare_tabu_months.py "תיקייה ישנה" "תיקייה חדשה"')
        print('דוגמה: python compare_tabu_months.py "נסחים מרץ 2026" "נסחים אפריל 2026"')
        sys.exit(1)

    old_folder = sys.argv[1]
    new_folder = sys.argv[2]

    print(f'📂 ישן: {old_folder}')
    print(f'📂 חדש: {new_folder}')
    print()

    print('🔍 מפענח את הנסחים הישנים...')
    old_data = parse_folder(old_folder)
    print(f'   {len(old_data)} חלקות')

    print('🔍 מפענח את הנסחים החדשים...')
    new_data = parse_folder(new_folder)
    print(f'   {len(new_data)} חלקות')

    print()
    print('🔄 השוואה...')
    diffs = {}
    all_parcels = sorted(set(old_data.keys()) | set(new_data.keys()), key=int)
    total_changes = 0
    for parcel in all_parcels:
        old_owners = old_data.get(parcel, {}).get('owners', [])
        new_owners = new_data.get(parcel, {}).get('owners', [])
        diff = compare_owners(old_owners, new_owners)
        n_changes = len(diff['added']) + len(diff['removed']) + len(diff['changed_area'])
        if n_changes > 0:
            diffs[parcel] = diff
            total_changes += n_changes
            print(f'  חלקה {parcel:>3}: +{len(diff["added"])} / -{len(diff["removed"])} / Δ{len(diff["changed_area"])}')

    print()
    print(f'📊 סה"כ שינויים: {total_changes}')

    if total_changes == 0:
        print('✅ אין שינויים — אין צורך בעדכון')
        return

    # שם הקובץ לפי שמות התיקיות
    old_short = old_folder.replace('נסחים ', '').replace(' ', '_')
    new_short = new_folder.replace('נסחים ', '').replace(' ', '_')
    os.makedirs('דוחות', exist_ok=True)
    out_path = f'דוחות/דוח_שינויים_{old_short}_ל{new_short}.xlsx'
    build_excel_report(diffs, out_path)

    print()
    print(f'📋 שלבים הבאים:')
    print(f'   1. פתח את {out_path}')
    print(f'   2. לכל שורה — סמן בעמודת "החלטה" אחת מ:')
    print(f'      • "להוסיף לחתומים" / "להוסיף לעסקה X"')
    print(f'      • "להסיר" / "לסמן is_active=FALSE"')
    print(f'      • "לעדכן שטח" / "להתעלם"')
    print(f'   3. תכתוב לי בצ\'אט "סיימתי את הדוח" ואני אבנה SQL')


if __name__ == '__main__':
    main()
