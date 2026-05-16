# -*- coding: utf-8 -*-
"""
פיתוח חד-פעמי - קריאת אקסל 'רשימת בעלי קרקע' והפקת רשימה לסקריפט SQL 95.
מבנה האקסל: עמודה 1 ריקה, עמודות 2-6 = שם, שטח, אחוז, חתימה (0/1), הערות
"""
import sys
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import openpyxl
from pathlib import Path

EXCEL_PATH = Path("../רשימת בעלי קרקע .xlsx")

wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
ws = wb.active

print("=" * 70)
print(f"sheet: {ws.title}, rows: {ws.max_row}, cols: {ws.max_column}")
print("=" * 70)

signed = []
not_signed = []
skipped = 0

for row_idx in range(4, ws.max_row + 1):
    name = ws.cell(row=row_idx, column=2).value
    area = ws.cell(row=row_idx, column=3).value
    pct = ws.cell(row=row_idx, column=4).value
    flag = ws.cell(row=row_idx, column=5).value
    notes = ws.cell(row=row_idx, column=6).value

    if name is None or str(name).strip() == '':
        skipped += 1
        continue

    name = str(name).strip()
    try:
        flag_int = int(flag) if flag is not None else 0
    except (ValueError, TypeError):
        flag_int = 0

    record = {
        'name': name,
        'area': area,
        'pct': pct,
        'flag': flag_int,
        'notes': str(notes).strip() if notes else None,
        'row': row_idx,
    }

    if flag_int == 1:
        signed.append(record)
    else:
        not_signed.append(record)

print(f"signed (flag=1): {len(signed)}")
print(f"NOT signed (flag=0): {len(not_signed)}")
print(f"total: {len(signed) + len(not_signed)}, skipped empty: {skipped}")
print("=" * 70)

# הצג דוגמאות
print("\n--- 10 signed (head) ---")
for r in signed[:10]:
    print(f"  row{r['row']}: {r['name']} | {r['area']} m^2 | flag={r['flag']}")

print("\n--- 10 NOT signed (head) ---")
for r in not_signed[:10]:
    print(f"  row{r['row']}: {r['name']} | {r['area']} m^2 | flag={r['flag']}")

# חיפוש גלזר מרים לאימות
print("\n--- search 'glazer' ---")
for r in signed + not_signed:
    if 'גלזר' in r['name'] or 'glazer' in r['name'].lower():
        print(f"  FOUND: row{r['row']}: {r['name']} | flag={r['flag']} | {r['area']} m^2 | notes: {r['notes']}")

# שמור לקובץ טקסט עבור הLLM
output_path = Path("../sql-scripts/_master_2015_names.txt")
with output_path.open("w", encoding="utf-8") as f:
    f.write("# Master 2015 landowners list\n")
    f.write(f"# Signed: {len(signed)}, NOT signed: {len(not_signed)}, Total: {len(signed)+len(not_signed)}\n\n")
    f.write("## SIGNED (flag=1)\n")
    for r in signed:
        f.write(f"{r['name']}\t{r['area']}\t{r['pct']}\t{r['notes'] or ''}\n")
    f.write("\n## NOT SIGNED (flag=0)\n")
    for r in not_signed:
        f.write(f"{r['name']}\t{r['area']}\t{r['pct']}\t{r['notes'] or ''}\n")

print(f"\nsaved to: {output_path}")
