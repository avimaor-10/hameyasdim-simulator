# -*- coding: utf-8 -*-
"""
בדיקה: כל השורות בגיליון "הסכמים שירי" - הצגת כל בעלי הקרקע
לפי חלקה אמיתית (col 26), כי לוין בוריס וליפ דביר בחלקה 7
"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl

wb = openpyxl.load_workbook("../עותק של 828579ניתוח קו כחול דוח חוקר בלי בית קברות.xlsx", data_only=True)
ws = wb["הסכמים שירי"]

print(f"sheet: {ws.title}, rows={ws.max_row}, cols={ws.max_column}\n")

# הדפס את כל ה-94 שורות עם החלקה האמיתית
print("ALL ROWS BY PARCEL:")
for row_idx in range(3, ws.max_row + 1):
    parcel = ws.cell(row=row_idx, column=26).value  # col 26 = חלקה
    main_name = ws.cell(row=row_idx, column=4).value
    main_id = ws.cell(row=row_idx, column=5).value
    heir_a = ws.cell(row=row_idx, column=7).value
    heir_b = ws.cell(row=row_idx, column=11).value
    signed = ws.cell(row=row_idx, column=23).value  # חתימה על ההסכם
    note = ws.cell(row=row_idx, column=32).value
    if main_name:
        print(f"  row{row_idx}: parcel={parcel}, main={main_name}({main_id}), heir_a={heir_a}, heir_b={heir_b}, signed_at={signed}")
        if note and ('לוין' in str(note) or 'ליפ' in str(note) or 'בוריס' in str(note)):
            print(f"    [NOTE] {note[:100]}")
