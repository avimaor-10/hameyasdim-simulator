# -*- coding: utf-8 -*-
"""
בונה סקריפט 95b ממוקד - רק 34 החריגים:
- 18 חתומים שלא ב-DB (שמות + שטח באקסל)
- 16 לא-חתומים שכן ב-DB (שמות + מה נמצא ב-DB)
"""
import sys, io, re
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import openpyxl
from pathlib import Path

EXCEL_PATH = Path("../רשימת בעלי קרקע .xlsx")
OUTPUT_PATH = Path("../sql-scripts/95b_diag_anomalies.sql")

wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
ws = wb.active

def clean_word(w):
    w = re.sub(r'[()\[\]"]', '', w)
    w = w.strip('-,. ')
    return w

def split_name(full_name):
    parts = re.split(r'\s+', full_name.strip())
    parts = [clean_word(p) for p in parts if clean_word(p)]
    parts = [p for p in parts if len(p) >= 2]
    parts.sort(key=len, reverse=True)
    return parts[:2] if len(parts) >= 2 else (parts + ['', ''])[:2]

records = []
for row_idx in range(4, ws.max_row + 1):
    name = ws.cell(row=row_idx, column=2).value
    area = ws.cell(row=row_idx, column=3).value
    flag = ws.cell(row=row_idx, column=5).value
    if name is None or str(name).strip() == '':
        continue
    name = str(name).strip()
    try:
        flag_int = int(flag) if flag is not None else 0
    except:
        flag_int = 0
    s1, s2 = split_name(name)
    records.append({'name': name, 'flag': flag_int, 'area': area, 's1': s1, 's2': s2})

def escape_sql(s):
    return s.replace("'", "''") if s else ''

cte_rows = ',\n'.join([
    f"    ('{escape_sql(r['name'])}', {r['flag']}, {r['area'] or 0}, '{escape_sql(r['s1'])}', '{escape_sql(r['s2'])}')"
    for r in records
])

sql = f"""-- ============================================================
-- 95b_diag_anomalies.sql  (16/05/2026)
-- ============================================================
-- 🎯 מטרה: לראות רק את 34 החריגים מ-95:
--   A) 18 חתומים באקסל שלא נמצאו ב-DB (אולי כתיב שונה)
--   B) 16 לא-חתומים באקסל שכן נמצאו ב-DB (אולי רכשנו מהם)
-- ============================================================


WITH excel_owners (excel_name, signed_flag, excel_area_sqm, search1, search2) AS (
  VALUES
{cte_rows}
), matched AS (
  SELECT
    e.excel_name,
    e.signed_flag,
    e.excel_area_sqm,
    COUNT(DISTINCT s.id_number) AS db_matches,
    STRING_AGG(DISTINCT s.owner_name, '; ' ORDER BY s.owner_name) AS db_names,
    STRING_AGG(DISTINCT s.id_number, '; ' ORDER BY s.id_number) AS db_ids
  FROM excel_owners e
  LEFT JOIN public.signed_owners s
    ON s.owner_name ILIKE '%' || e.search1 || '%'
   AND s.owner_name ILIKE '%' || e.search2 || '%'
   AND s.is_active = TRUE
  GROUP BY e.excel_name, e.signed_flag, e.excel_area_sqm
)
-- שאילתה A: 18 חתומים שלא נמצאו ב-DB
SELECT
  '⚠ A. חתם באקסל - לא ב-DB' AS "סוג",
  excel_name                 AS "שם באקסל",
  ROUND(excel_area_sqm::numeric, 1) AS "שטח באקסל",
  NULL                       AS "שמות ב-DB",
  NULL                       AS "ת״ז ב-DB"
FROM matched
WHERE signed_flag = 1 AND db_matches = 0

UNION ALL

-- שאילתה B: 16 לא-חתומים שכן נמצאו ב-DB
SELECT
  '⚠ B. לא חתם באקסל - כן ב-DB',
  excel_name,
  ROUND(excel_area_sqm::numeric, 1),
  db_names,
  db_ids
FROM matched
WHERE signed_flag = 0 AND db_matches > 0

ORDER BY "סוג", "שטח באקסל" DESC NULLS LAST;
"""

OUTPUT_PATH.write_text(sql, encoding='utf-8')
print(f"saved: {OUTPUT_PATH}")
