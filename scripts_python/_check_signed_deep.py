# -*- coding: utf-8 -*-
"""
בדיקה עמוקה: הדפסת כל שמות בעלי הקרקע מ-2 הגיליונות הקריטיים.
1. רשימת בעלי קרקע חתומים על הסכם ניהול.xlsx - גיליון1 (201 שורות)
2. עותק של 828579ניתוח קו כחול דוח חוקר בלי בית קברות.xlsx - גיליון "הסכמים שירי" (94 שורות)
"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl
from pathlib import Path

# ========== קובץ 1 ==========
print("="*70)
print("FILE 1: רשימת בעלי קרקע חתומים על הסכם ניהול.xlsx")
print("="*70)
wb1 = openpyxl.load_workbook("../רשימת בעלי קרקע חתומים על הסכם ניהול.xlsx", data_only=True)
ws1 = wb1["גיליון1"]

# שורה 2 = כותרות, שורה 3+ = נתונים
print("Columns:")
for col_idx in range(1, ws1.max_column + 1):
    print(f"  col{col_idx}: {ws1.cell(row=2, column=col_idx).value}")

print(f"\nTotal rows: {ws1.max_row}")
print("\n--- All names (col 4 = שם בעל הקרקע) ---")
for row_idx in range(3, ws1.max_row + 1):
    name = ws1.cell(row=row_idx, column=4).value
    id_num = ws1.cell(row=row_idx, column=5).value
    parcel = ws1.cell(row=row_idx, column=3).value
    if name:
        print(f"  row{row_idx}: parcel={parcel} | name={name} | id={id_num}")


# ========== קובץ 2 ==========
print("\n\n" + "="*70)
print("FILE 2: עותק של 828579ניתוח קו כחול - גיליון 'הסכמים שירי'")
print("="*70)
wb2 = openpyxl.load_workbook("../עותק של 828579ניתוח קו כחול דוח חוקר בלי בית קברות.xlsx", data_only=True)
ws2 = wb2["הסכמים שירי"]

# שורה 2 = כותרות
print("Columns:")
for col_idx in range(1, min(ws2.max_column + 1, 15)):
    print(f"  col{col_idx}: {ws2.cell(row=2, column=col_idx).value}")

print(f"\nTotal rows: {ws2.max_row}")
print("\n--- All names (col 4 = שם בעל קרקע) ---")
for row_idx in range(3, ws2.max_row + 1):
    name = ws2.cell(row=row_idx, column=4).value
    id_num = ws2.cell(row=row_idx, column=5).value
    parcel = ws2.cell(row=row_idx, column=3).value
    heir_a = ws2.cell(row=row_idx, column=7).value
    if name:
        print(f"  row{row_idx}: parcel={parcel} | name={name} | id={id_num} | heir_a={heir_a}")
