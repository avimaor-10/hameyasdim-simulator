# -*- coding: utf-8 -*-
"""
בדיקה של 2 אקסלים נוספים לחיפוש 9 הנותרים:
1. רשימת בעלי קרקע חתומים על הסכם ניהול.xlsx
2. עותק של 828579ניתוח קו כחול דוח חוקר בלי בית קברות.xlsx
"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl
from pathlib import Path

EXCELS = [
    Path("../רשימת בעלי קרקע חתומים על הסכם ניהול.xlsx"),
    Path("../עותק של 828579ניתוח קו כחול דוח חוקר בלי בית קברות.xlsx"),
]

# 9 השמות שלא אומתו עדיין
SEARCH_NAMES = [
    "דרורי אליהו", "דרורי מיכל",
    "סטולר אלה", "סטולר אורה",
    "פלדמן אהרון",
    "שוורצברג אריאלה",
    "לוין בוריס", "בוריס לוין",
    "ליפ דביר", "ליפ אייל",
    "גלזר מרים",
    "דניאלי סמדר",
]

for excel_path in EXCELS:
    if not excel_path.exists():
        print(f"\n=== FILE NOT FOUND: {excel_path} ===\n")
        continue
    print(f"\n{'='*70}")
    print(f"FILE: {excel_path.name}")
    print(f"{'='*70}")
    wb = openpyxl.load_workbook(excel_path, data_only=True)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n--- sheet: {sheet_name} (rows={ws.max_row}, cols={ws.max_column}) ---")

        # הצג שורה 1-3 כדי לראות מבנה
        for row_idx in range(1, min(4, ws.max_row + 1)):
            row_data = []
            for col_idx in range(1, min(ws.max_column + 1, 12)):
                val = ws.cell(row=row_idx, column=col_idx).value
                row_data.append(str(val)[:30] if val else "_")
            print(f"  row{row_idx}: {' | '.join(row_data)}")

        # חיפוש בכל הגיליון - לכל שם מהרשימה
        print(f"\n  matches in this sheet:")
        for search in SEARCH_NAMES:
            found_in = []
            for row in ws.iter_rows(values_only=True):
                for cell in row:
                    if cell is None:
                        continue
                    cell_str = str(cell)
                    if search in cell_str:
                        found_in.append(cell_str[:80])
                        break
            if found_in:
                # הסר כפילויות
                unique = list(dict.fromkeys(found_in))[:3]
                for match in unique:
                    print(f"    [{search}] -> {match}")
