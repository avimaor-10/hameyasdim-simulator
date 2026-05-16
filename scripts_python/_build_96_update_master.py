# -*- coding: utf-8 -*-
"""
בונה סקריפט 96 - UPDATE master_2015_status='verified' לכל רשומה ב-DB
שיש לה התאמה לאחד מ-199 השמות באקסל.
"""
import sys, io, re
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import openpyxl
from pathlib import Path

EXCEL_PATH = Path("../רשימת בעלי קרקע .xlsx")
OUTPUT_PATH = Path("../sql-scripts/96_update_master_2015_status.sql")

wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
ws = wb.active

def clean_word(w):
    w = re.sub(r'[()\[\]"]', '', w)
    w = w.strip('-,. ')
    return w

def split_name(full_name):
    parts = re.split(r'\s+', full_name.strip())
    parts = [clean_word(p) for p in parts if clean_word(p)]
    parts = [p for p in parts if len(p) >= 2]
    parts.sort(key=len, reverse=True)
    return parts[:2] if len(parts) >= 2 else (parts + ['', ''])[:2]

records = []
for row_idx in range(4, ws.max_row + 1):
    name = ws.cell(row=row_idx, column=2).value
    area = ws.cell(row=row_idx, column=3).value
    flag = ws.cell(row=row_idx, column=5).value
    if name is None or str(name).strip() == '':
        continue
    name = str(name).strip()
    try:
        flag_int = int(flag) if flag is not None else 0
    except:
        flag_int = 0
    s1, s2 = split_name(name)
    records.append({'name': name, 'flag': flag_int, 'area': area, 's1': s1, 's2': s2})

def escape_sql(s):
    return s.replace("'", "''") if s else ''

# רק רשומות עם 2 מילות חיפוש תקפות
valid = [r for r in records if r['s1'] and r['s2']]
print(f"records with valid 2 search words: {len(valid)} (of {len(records)})")

cte_rows = ',\n'.join([
    f"    ('{escape_sql(r['name'])}', {r['flag']}, '{escape_sql(r['s1'])}', '{escape_sql(r['s2'])}')"
    for r in valid
])

sql = f"""-- ============================================================
-- 96_update_master_2015_status.sql  (16/05/2026)
-- ============================================================
-- 🎯 מטרה: סימון master_2015_status='verified' לכל רשומה ב-DB
--          שמותאמת לבעלים באקסל 'רשימת בעלי קרקע' (חתום או לא).
--
-- 📐 לוגיקה:
--   • כל רשומה ב-DB שמותאמת לבעלים באקסל - verified
--     - אם חתם (flag=1): השרשור אומת ישירות
--     - אם לא חתם (flag=0): הוא ב-DB דרך שרשור (יורש/רכישה)
--       כלומר - אומת דרך מי שלפניו
--
--   • COALESCE - לא דורס 'verified' קיים. אם כבר verified, נשאר.
--   • את master_2015_notes ממלא רק אם NULL (לא דורס).
--
-- 🔒 בטוח להריצה - רק מעלה רמת אישור, לא מורידה.
-- ============================================================


BEGIN;


-- ============================================================
-- שאילתה 1: לפני - פיזור master_2015_status
-- ============================================================
SELECT '📸 לפני 96' AS "שלב",
  COALESCE(master_2015_status, '(NULL)') AS "סטטוס",
  COUNT(*) AS "כמות"
FROM public.signed_owners
WHERE is_active = TRUE
GROUP BY master_2015_status
ORDER BY COUNT(*) DESC;


-- ============================================================
-- צעד A: UPDATE master_2015_status='verified' לכל מי שמותאם לאקסל
-- ============================================================
WITH excel_owners (excel_name, signed_flag, search1, search2) AS (
  VALUES
{cte_rows}
), to_verify AS (
  SELECT DISTINCT s.id, e.excel_name, e.signed_flag
  FROM excel_owners e
  JOIN public.signed_owners s
    ON s.owner_name ILIKE '%' || e.search1 || '%'
   AND s.owner_name ILIKE '%' || e.search2 || '%'
   AND s.is_active = TRUE
)
UPDATE public.signed_owners s
SET
  master_2015_status = 'verified',
  master_2015_notes  = COALESCE(s.master_2015_notes,
    'מותאם לאקסל מאסטר 2015: ' || tv.excel_name ||
    CASE tv.signed_flag WHEN 1 THEN ' (חתם)' ELSE ' (לא חתם - דרך שרשור)' END
  )
FROM to_verify tv
WHERE s.id = tv.id;


-- ============================================================
-- שאילתה 2: אחרי - פיזור master_2015_status
-- ============================================================
SELECT '✅ אחרי 96' AS "שלב",
  COALESCE(master_2015_status, '(NULL)') AS "סטטוס",
  COUNT(*) AS "כמות"
FROM public.signed_owners
WHERE is_active = TRUE
GROUP BY master_2015_status
ORDER BY COUNT(*) DESC;


-- ============================================================
-- שאילתה 3: דוגמה - 20 רשומות שעודכנו (לאימות)
-- ============================================================
SELECT
  s.owner_name AS "שם",
  s.id_number  AS "ת״ז",
  s.parcel     AS "חלקה",
  s.master_2015_status AS "סטטוס",
  LEFT(s.master_2015_notes, 60) AS "הערה (60 תווים)"
FROM public.signed_owners s
WHERE s.master_2015_status = 'verified'
  AND s.is_active = TRUE
ORDER BY s.owner_name
LIMIT 20;


COMMIT;


-- ============================================================
-- 📋 צפי תוצאות:
--   "אחרי 96" צריך להראות:
--     • verified: ~80-100 רשומות (היו ~24 לפני)
--     • NULL/(NULL): פחות מאשר לפני
--
--   אם עלייה קטנה - יכול להיות שחלק כבר היו verified.
-- ============================================================
"""

OUTPUT_PATH.write_text(sql, encoding='utf-8')
print(f"saved: {OUTPUT_PATH}")
print(f"file size: {OUTPUT_PATH.stat().st_size} bytes")
