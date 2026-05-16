# -*- coding: utf-8 -*-
"""
בדיקה מדויקת של גלזר מרים בקובץ "רשימת בעלי קרקע חתומים על הסכם ניהול.xlsx"
+ עמודות 6-11 (שטח, סטטוס, פרטים, הערה)
"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl
from pathlib import Path

wb = openpyxl.load_workbook("../רשימת בעלי קרקע חתומים על הסכם ניהול.xlsx", data_only=True)
ws = wb["גיליון1"]

# חפש את גלזר וגם דניאלי סמדר כדי להשוות
SEARCHES = ['גלזר', 'דניאלי סמדר', 'סמדר', 'לוין בוריס', 'בוריס', 'ליפ', 'דרורי', 'סטולר', 'פלדמן אהרון', 'שוורצברג']

# מצא את כל השורות שמכילות שמות אלה
for search in SEARCHES:
    print(f"\n--- search: '{search}' ---")
    for row_idx in range(3, ws.max_row + 1):
        name = ws.cell(row=row_idx, column=4).value
        if name is None:
            continue
        if search in str(name):
            parcel = ws.cell(row=row_idx, column=3).value
            id_num = ws.cell(row=row_idx, column=5).value
            area_plan = ws.cell(row=row_idx, column=6).value
            area_deal = ws.cell(row=row_idx, column=7).value
            area_united = ws.cell(row=row_idx, column=8).value
            status = ws.cell(row=row_idx, column=9).value
            details = ws.cell(row=row_idx, column=10).value
            note = ws.cell(row=row_idx, column=11).value
            print(f"  row{row_idx}: parcel={parcel}, name={name}, id={id_num}")
            print(f"    area_plan={area_plan}, area_deal={area_deal}, area_united={area_united}")
            print(f"    status={status}")
            print(f"    details={details}")
            print(f"    note={note}")
