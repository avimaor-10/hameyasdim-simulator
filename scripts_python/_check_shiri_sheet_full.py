# -*- coding: utf-8 -*-
"""
הדפסת כל 94 השורות + 39 העמודות של גיליון "הסכמים שירי"
ולחפש את לוין, ליפ, בוריס בכל מקום
"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl
from pathlib import Path

wb = openpyxl.load_workbook("../עותק של 828579ניתוח קו כחול דוח חוקר בלי בית קברות.xlsx", data_only=True)
ws = wb["הסכמים שירי"]

print(f"sheet: {ws.title}, rows={ws.max_row}, cols={ws.max_column}")
print()

# הדפס כותרות
print("=" * 70)
print("HEADERS (row 2):")
for col_idx in range(1, ws.max_column + 1):
    print(f"  col{col_idx}: {ws.cell(row=2, column=col_idx).value}")

# חיפוש מקיף — בכל עמודה
SEARCHES = ['לוין', 'ליפ', 'בוריס', 'דרורי', 'סטולר', 'פלדמן אהרון', 'שוורצברג', 'דניאלי', 'גלזר']

print("\n" + "=" * 70)
print("SEARCH ALL CELLS IN ALL ROWS:")
for search in SEARCHES:
    print(f"\n--- '{search}' ---")
    found = False
    for row_idx in range(1, ws.max_row + 1):
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx).value
            if cell is None:
                continue
            if search in str(cell):
                # הדפס את כל השורה
                parcel = ws.cell(row=row_idx, column=3).value
                main_name = ws.cell(row=row_idx, column=4).value
                main_id = ws.cell(row=row_idx, column=5).value
                heir_a_name = ws.cell(row=row_idx, column=7).value
                heir_a_id = ws.cell(row=row_idx, column=8).value
                heir_b_name = ws.cell(row=row_idx, column=11).value
                heir_b_id = ws.cell(row=row_idx, column=12).value
                heir_c_name = ws.cell(row=row_idx, column=15).value if ws.max_column >= 15 else None
                heir_c_id = ws.cell(row=row_idx, column=16).value if ws.max_column >= 16 else None
                # תאריך חתימה (אם יש)
                signed_at_col = None
                for c in range(1, ws.max_column + 1):
                    val = ws.cell(row=row_idx, column=c).value
                    if val and ('חתימה על' in str(val) or '/' in str(val)[:10]):
                        signed_at_col = (c, val)
                print(f"  row{row_idx} col{col_idx}: '{str(cell)[:60]}'")
                print(f"    parcel={parcel}, main_name={main_name} (id={main_id})")
                print(f"    heir_a={heir_a_name} (id={heir_a_id})")
                print(f"    heir_b={heir_b_name} (id={heir_b_id})")
                print(f"    heir_c={heir_c_name} (id={heir_c_id})")
                found = True
                break  # אחד לשורה
    if not found:
        print(f"  NOT FOUND")
