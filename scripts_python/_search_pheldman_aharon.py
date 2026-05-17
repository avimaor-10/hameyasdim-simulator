# -*- coding: utf-8 -*-
"""
חיפוש פלדמן אהרון בכל אקסלי האימות - לראות אם הוא חתום.
מחפש לפי שם ולפי כל וריאציה של ת"ז: 156138 / 001561380 / 000156138 / 1561380
"""
import sys, io, os
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import openpyxl
from openpyxl.utils.exceptions import InvalidFileException

ROOT = r"C:\Users\User2\Desktop\סימולטור עסקת קומבינציה וטבלאות ההקצעה"

# אקסלים מקוטלגים לבדיקה
EXCEL_FILES = [
    # ראשי - רשימות חתימה והסכמים
    "דוח_הסכמי_ניהול_ואיחוד_וחלוקה_גוש_3852_v13.xlsx",
    "רשימת בעלי קרקע שחתומים על הסכם מקור.xlsx",
    "רשימת בעלי קרקע חתומים על הסכם ניהול.xlsx",
    "רשימת בעלי קרקע מקור ופרטי התקשרות.xlsx",
    "רשימת בעלי קרקע .xlsx",
    "812576בעלים חדשים.xlsx",
    "824336_אחזקה לפי בעלות - נכון ליום 2.2.xlsx",
    # הצלבה ואימות
    "דוח_אי_התאמה_בעלים.xlsx",
    "דוח_הצלבה_טאבו_מרץ_2026.xlsx",
    "טבלת בעלי זכויות מהמנדיי מאומת נסח טאבו מרץ 26.xlsx",
]

# וריאציות של ת"ז וזיהוי
ID_VARIANTS = ['156138', '001561380', '000156138', '1561380', '0001561380']
NAME_PATTERNS = ['אהרון', 'אהרן']  # פלדמן אהרון/אהרן
LAST_NAME = 'פלדמן'


def cell_to_str(val):
    """המרת ערך תא לטקסט בטוח לחיפוש"""
    if val is None:
        return ''
    if isinstance(val, (int, float)):
        # מספרים יכולים להיות ת"ז עם 0 בהתחלה שאבד
        return str(int(val)) if val == int(val) else str(val)
    return str(val)


def find_matches(file_path):
    """חיפוש בכל הגיליונות של קובץ אקסל"""
    results = []
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
    except (InvalidFileException, OSError) as e:
        return [('ERROR', '', '', f'cannot open: {e}')]

    for sheet_name in wb.sheetnames:
        try:
            ws = wb[sheet_name]
        except Exception:
            continue

        for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            row_text_parts = [cell_to_str(v) for v in row]
            row_text = ' | '.join(row_text_parts)

            # האם השורה מכילה פלדמן + אהרון/אהרן
            has_feldman = LAST_NAME in row_text
            has_aharon = any(p in row_text for p in NAME_PATTERNS)
            has_id = any(v in row_text for v in ID_VARIANTS)

            if (has_feldman and has_aharon) or has_id:
                results.append((sheet_name, row_idx, row_text[:300], ''))

    wb.close()
    return results


print("="*100)
print("חיפוש פלדמן אהרון/אהרן (ת'ז 156138 / 001561380 / 000156138) בכל האקסלים")
print("="*100)
print()

for filename in EXCEL_FILES:
    path = os.path.join(ROOT, filename)
    if not os.path.exists(path):
        print(f"❌ לא נמצא: {filename}")
        continue

    print(f"\n📂 {filename}")
    print("-" * 100)

    matches = find_matches(path)
    if not matches:
        print("   ⚪ לא נמצאו רשומות")
        continue

    for sheet_name, row_num, text, err in matches:
        if err:
            print(f"   ⚠ {err}")
        else:
            print(f"   📄 [{sheet_name}] שורה {row_num}: {text}")

print("\n" + "="*100)
print("סיום החיפוש")
print("="*100)
