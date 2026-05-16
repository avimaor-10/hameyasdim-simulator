# -*- coding: utf-8 -*-
"""
בונה את סקריפט SQL 95 (אבחון) מתוך אקסל 'רשימת בעלי קרקע'.
הסקריפט יראה לכל אחד מ-199 השמות באקסל - האם הוא מותאם ל-owner_name ב-DB.
"""
import sys, io, re
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import openpyxl
from pathlib import Path

EXCEL_PATH = Path("../רשימת בעלי קרקע .xlsx")
OUTPUT_PATH = Path("../sql-scripts/95_diag_master_2015_match.sql")

wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
ws = wb.active

def clean_word(w):
    """ניקוי מילה - הסרת סוגריים, מקפים, גרשיים"""
    w = re.sub(r'[()\[\]"]', '', w)
    w = w.strip('-,. ')
    return w

def split_name(full_name):
    """מחזיר את 2 המילים הארוכות (הכי מבדילות) מהשם"""
    parts = re.split(r'\s+', full_name.strip())
    parts = [clean_word(p) for p in parts if clean_word(p)]
    parts = [p for p in parts if len(p) >= 2]  # מילים קצרות לא יעילות
    # מיון לפי אורך יורד
    parts.sort(key=len, reverse=True)
    return parts[:2] if len(parts) >= 2 else (parts + ['', ''])[:2]

records = []
for row_idx in range(4, ws.max_row + 1):
    name = ws.cell(row=row_idx, column=2).value
    area = ws.cell(row=row_idx, column=3).value
    flag = ws.cell(row=row_idx, column=5).value

    if name is None or str(name).strip() == '':
        continue

    name = str(name).strip()
    try:
        flag_int = int(flag) if flag is not None else 0
    except (ValueError, TypeError):
        flag_int = 0

    s1, s2 = split_name(name)
    records.append({
        'name': name,
        'flag': flag_int,
        'area': area,
        's1': s1,
        's2': s2,
    })

print(f"records: {len(records)}")
print(f"signed: {sum(1 for r in records if r['flag']==1)}")
print(f"not signed: {sum(1 for r in records if r['flag']==0)}")

# בנה את ה-CTE
def escape_sql(s):
    return s.replace("'", "''") if s else ''

cte_rows = []
for r in records:
    line = f"    ('{escape_sql(r['name'])}', {r['flag']}, {r['area'] or 0}, '{escape_sql(r['s1'])}', '{escape_sql(r['s2'])}')"
    cte_rows.append(line)

cte_block = ',\n'.join(cte_rows)

# בנה את הסקריפט
sql = f"""-- ============================================================
-- 95_diag_master_2015_match.sql  (16/05/2026)
-- ============================================================
-- 🎯 מטרה: סקריפט אבחנתי בלבד (SELECT, לא UPDATE) -
--          הצלבת כל 199 השמות באקסל 'רשימת בעלי קרקע' אל
--          ה-DB, וזיהוי איזה בעלים מותאמים ואיזה לא.
--
-- 📐 שיטה:
--   • CTE עם 199 שורות מהאקסל (שם, flag, שטח, 2 מילות חיפוש)
--   • LEFT JOIN עם signed_owners לפי 2 מילות חיפוש
--   • החזרה: לכל שם באקסל - האם נמצא ב-DB? כמה התאמות?
-- ============================================================


WITH excel_owners (excel_name, signed_flag, excel_area_sqm, search1, search2) AS (
  VALUES
{cte_block}
)
SELECT
  e.excel_name                                    AS "שם באקסל",
  CASE e.signed_flag WHEN 1 THEN '✅ חתם' ELSE '❌ לא חתם' END AS "סטטוס באקסל",
  ROUND(e.excel_area_sqm::numeric, 1)             AS "שטח באקסל",
  COUNT(DISTINCT s.id_number)                     AS "התאמות ב-DB",
  STRING_AGG(DISTINCT s.owner_name, '; ')         AS "שמות שנמצאו ב-DB"
FROM excel_owners e
LEFT JOIN public.signed_owners s
  ON s.owner_name ILIKE '%' || e.search1 || '%'
 AND s.owner_name ILIKE '%' || e.search2 || '%'
 AND s.is_active = TRUE
GROUP BY e.excel_name, e.signed_flag, e.excel_area_sqm
ORDER BY e.signed_flag DESC, "התאמות ב-DB" ASC, "שטח באקסל" DESC NULLS LAST;


-- ============================================================
-- שאילתה 2: סיכום מהיר - כמה מתוך 199 נמצאו ב-DB
-- ============================================================
WITH excel_owners (excel_name, signed_flag, search1, search2) AS (
  VALUES
{','.join([chr(10) + f'    (' + chr(39) + escape_sql(r['name']) + chr(39) + ', ' + str(r['flag']) + ', ' + chr(39) + escape_sql(r['s1']) + chr(39) + ', ' + chr(39) + escape_sql(r['s2']) + chr(39) + ')' for r in records])}
), matched AS (
  SELECT
    e.excel_name,
    e.signed_flag,
    COUNT(DISTINCT s.id_number) AS db_matches
  FROM excel_owners e
  LEFT JOIN public.signed_owners s
    ON s.owner_name ILIKE '%' || e.search1 || '%'
   AND s.owner_name ILIKE '%' || e.search2 || '%'
   AND s.is_active = TRUE
  GROUP BY e.excel_name, e.signed_flag
)
SELECT
  CASE signed_flag WHEN 1 THEN '✅ חתמו (flag=1)' ELSE '❌ לא חתמו (flag=0)' END AS "קבוצה",
  COUNT(*)                                            AS "סה''כ באקסל",
  COUNT(*) FILTER (WHERE db_matches > 0)              AS "נמצאו ב-DB",
  COUNT(*) FILTER (WHERE db_matches = 0)              AS "לא נמצאו ב-DB"
FROM matched
GROUP BY signed_flag
ORDER BY signed_flag DESC;


-- ============================================================
-- 📋 צפי תוצאות:
--   שאילתה 1: 199 שורות, כל שם באקסל עם כמות ההתאמות שלו ב-DB.
--             חתומים אמורים להיות עם 1+ התאמות (יחד עם יורשים).
--             לא-חתומים אמורים להיות עם 0 התאמות (לא נכנסו ל-DB).
--
--   שאילתה 2: סיכום מספרי - מי נמצא ומי לא.
-- ============================================================
"""

OUTPUT_PATH.write_text(sql, encoding='utf-8')
print(f"saved: {OUTPUT_PATH}")
print(f"file size: {OUTPUT_PATH.stat().st_size} bytes")
