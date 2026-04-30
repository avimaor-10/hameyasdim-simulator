"""בדיקה: לכל חלקה — חתומים + צד ג' = שטח חלקה? והאם עסקאות יוצרות כפילות?"""
import json
import sys
import openpyxl
import re

sys.stdout.reconfigure(encoding='utf-8')

def normalize_id(s):
    return str(s or '').strip().lstrip('0')

# טען נתוני טאבו
with open('assets/tabu_owners.json', encoding='utf-8') as f:
    tabu = json.load(f)

# טען חתומים מהאקסל המקורי v13 (נחשב לסופרבייס בקירוב)
wb = openpyxl.load_workbook('דוח_הסכמי_ניהול_ואיחוד_וחלוקה_גוש_3852_v13.xlsx', data_only=True)
ws = wb['בעלים — חתומים על הסכם ניהול']
signed_by_parcel = {}
for row in ws.iter_rows(min_row=3, values_only=True):
    if not row[2] or row[2] == 'מס"ד': continue
    try:
        parcel = int(row[2])
        id_ = normalize_id(row[4])
        area = float(row[7] or 0)
        if parcel not in signed_by_parcel:
            signed_by_parcel[parcel] = []
        signed_by_parcel[parcel].append({'id': id_, 'name': row[3], 'area': area})
    except (ValueError, TypeError):
        continue

# טען עסקאות מאקסל המקור v13
ws = wb['עסקאות שותפויות חנן מור']
deals_by_parcel = {}
for row in ws.iter_rows(min_row=4, values_only=True):
    if not row[0] or not isinstance(row[2], (int, float)) or row[2] is None: continue
    try:
        parcel = int(row[2])
        area_raw = row[3]
        area = float(str(area_raw).replace(',', '')) if area_raw else 0
        if parcel not in deals_by_parcel:
            deals_by_parcel[parcel] = []
        deals_by_parcel[parcel].append({'name': row[1], 'area': area})
    except (ValueError, TypeError):
        continue

# בדיקה לכל חלקה
print(f'{"חלקה":<6} {"שטח":<8} {"טאבו":<10} {"חתומים":<10} {"צד ג":<10} {"עסקאות":<10} {"סטטוס"}')
print('-' * 90)

issues = 0
for parcel_str, parcel_data in sorted(tabu.items(), key=lambda x: int(x[0])):
    parcel = int(parcel_str)
    parcel_area = parcel_data['parcel_area']
    tabu_total = parcel_data['owners_total_area']

    # ת.ז. של חתומים
    signed_ids = set(normalize_id(s['id']) for s in signed_by_parcel.get(parcel, []))
    signed_area_in_tabu = sum(o['area_sqm'] for o in parcel_data['owners']
                              if normalize_id(o['id']) in signed_ids)
    third_party_area = tabu_total - signed_area_in_tabu

    # עסקאות
    deals_area = sum(d['area'] for d in deals_by_parcel.get(parcel, []))

    # סטטוס
    sum_check = signed_area_in_tabu + third_party_area
    diff = abs(sum_check - tabu_total)
    overlap = signed_area_in_tabu + third_party_area + deals_area - parcel_area

    if abs(diff) < 1 and abs(overlap) < 5:
        status = '✓ תקין'
    elif overlap > 5:
        status = f'⚠ עודף עסקאות {overlap:.0f}'
        issues += 1
    elif overlap < -5:
        status = f'⚠ חסר {-overlap:.0f}'
        issues += 1
    else:
        status = '✓ תקין'

    print(f'{parcel:<6} {parcel_area:<8} {tabu_total:<10.0f} {signed_area_in_tabu:<10.0f} {third_party_area:<10.0f} {deals_area:<10.0f} {status}')

print()
print(f'סה"כ חלקות עם בעיה: {issues}')
